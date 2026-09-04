# -*- coding: utf-8 -*-
"""
Genera el SQL para importar pacientes desde el Excel de la clinica.

Solo INSERTA los que faltan (ON CONFLICT (dni) DO NOTHING): nunca modifica ni borra
un paciente existente, asi que las citas/pagos que ya apuntan a ellos no se tocan.

Uso:  python generar_import_pacientes.py <archivo.xlsx> <salida.sql> [--correo-sintetico]
"""
import sys, re, unicodedata, datetime, collections
import openpyxl

# Particulas que forman parte del apellido cuando aparecen pegadas a el
# ("DE LA CRUZ", "DEL AGUILA"). El caso ambiguo es cuando forman parte del NOMBRE
# ("GABRIELA DEL PILAR"): ahi la heuristica se apoya en que el nombre de pila va antes.
PARTICULAS = {'DE', 'DEL', 'LA', 'LAS', 'LOS', 'Y', 'DA', 'DI', 'VAN', 'VON', "D'"}

# Filas que claramente no son un paciente real.
BASURA = {'DUPLICADO', 'N.A.', 'NA', 'SIN NOMBRE', 'XXX', '.', '-'}


def limpiar(v):
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def normalizar_espacios(s):
    return re.sub(r'\s+', ' ', s).strip()


def partir_nombre(completo):
    """
    Convencion peruana: NOMBRE(S) APELLIDO_PATERNO APELLIDO_MATERNO.
    Devuelve (nombre, apellido, ambiguo) — `ambiguo` marca los casos donde la
    heuristica pudo equivocarse, para poder revisarlos despues.

    Reglas, en orden:
      - 1 palabra  -> todo va a apellido (nombre no puede ir vacio: es NOT NULL).
      - 2 palabras -> una y una.
      - "... APELLIDO1 APELLIDO2 DE ESPOSO" (apellido de casada): los ultimos 4.
      - "... DE LA CRUZ PRADO": se absorben las particulas que preceden al bloque.
      - resto      -> los ultimos 2 son apellidos.
    """
    p = normalizar_espacios(completo).upper().split()
    if not p:
        return ('', '', True)
    if len(p) == 1:
        return ('(SIN NOMBRE)', p[0], True)
    if len(p) == 2:
        return (p[0], p[1], False)

    ambiguo = False

    # Apellido de casada: "MARIA CUEVA HUAMAN DE ALVARADO" -> apellido "CUEVA HUAMAN DE ALVARADO"
    if p[-2] in PARTICULAS and len(p) >= 5:
        corte = len(p) - 4
    else:
        corte = len(p) - 2
        # "GLADYS RUTH DE LA CRUZ PRADO" -> el bloque crece hacia atras mientras haya particulas
        while corte > 1 and p[corte - 1] in PARTICULAS:
            corte -= 1
            ambiguo = True

    if corte < 1:            # no quedaria nombre: se deja al menos la primera palabra
        corte = 1
        ambiguo = True

    nombre = ' '.join(p[:corte])
    apellido = ' '.join(p[corte:])
    if len(p) >= 5:
        ambiguo = True       # nombres largos: siempre conviene una revision humana
    return (nombre, apellido, ambiguo)


def escapar(s):
    return s.replace("'", "''")


def sql_txt(s):
    return "NULL" if not s else "'" + escapar(s) + "'"


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    entrada, salida = sys.argv[1], sys.argv[2]
    correo_sintetico = '--correo-sintetico' in sys.argv

    wb = openpyxl.load_workbook(entrada, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(min_row=2, values_only=True))

    vistos = set()
    registros = []
    stats = collections.Counter()
    ambiguos = []

    for f in filas:
        fuente, _fing, nombre_completo, dni, tel, fnac, sexo = (list(f) + [None] * 7)[:7]
        dni = limpiar(dni).upper()
        nombre_completo = normalizar_espacios(limpiar(nombre_completo)).upper()

        if not dni or not nombre_completo:
            stats['sin_dni_o_nombre'] += 1
            continue
        if nombre_completo in BASURA or dni in BASURA or ' ' in dni:
            stats['basura'] += 1
            continue
        if dni in vistos:
            stats['duplicado_en_excel'] += 1
            continue
        vistos.add(dni)

        nombre, apellido, ambiguo = partir_nombre(nombre_completo)
        if ambiguo:
            ambiguos.append((dni, nombre_completo, nombre, apellido))

        telefono = re.sub(r'\D', '', limpiar(tel))[:20] or ''
        fecha_nac = limpiar(fnac) or ''
        sexo_txt = limpiar(sexo).upper()
        fuente_txt = limpiar(fuente).upper()

        # El Excel trae SEXO y FUENTE, que no tienen columna propia en la tabla:
        # se conservan en notas para no perder el dato.
        notas = ' · '.join(x for x in [
            'Importado del Excel del 03-09-2026',
            ('Fuente: ' + fuente_txt) if fuente_txt else '',
            ('Sexo: ' + sexo_txt) if sexo_txt else '',
        ] if x)

        correo = ''
        if correo_sintetico:
            correo = 'paciente.%s@sincorreo.local' % re.sub(r'[^A-Za-z0-9]', '', dni).lower()

        registros.append((nombre[:100], apellido[:100], dni[:20], telefono, correo, fecha_nac, notas))
        stats['listos'] += 1

    with open(salida, 'w', encoding='utf-8') as out:
        w = out.write
        w("-- ══════════════════════════════════════════════════════════════════════════\n")
        w("--  IMPORTACION DE PACIENTES — Excel del 03-09-2026\n")
        w("--\n")
        w("--  SOLO AGREGA los que faltan. ON CONFLICT (dni) DO NOTHING: si el DNI ya\n")
        w("--  existe, la fila se ignora por completo — no se pisa ningun dato del\n")
        w("--  paciente que ya estaba, asi que sus citas, pagos y paquetes quedan intactos.\n")
        w("--\n")
        w("--  Es idempotente: correrlo dos veces no duplica nada.\n")
        w("--\n")
        w("--  IRREVERSIBLE en el sentido de que agrega %d filas. Backup antes:\n" % len(registros))
        w("--    docker compose exec -T db pg_dump -U postgres -d BDClinicaSAAS -F c > ~/backups/antes-import.dump\n")
        w("-- ══════════════════════════════════════════════════════════════════════════\n\n")
        w("BEGIN;\n\n")
        w("SELECT 'ANTES' AS momento, count(*) AS pacientes FROM pacientes;\n\n")
        w("-- Tabla temporal: se cargan todas las filas del Excel y luego se insertan de una,\n")
        w("-- lo que permite reportar cuantas entraron y cuantas ya existian.\n")
        w("CREATE TEMP TABLE _import_pacientes (\n")
        w("    nombre           VARCHAR(100),\n")
        w("    apellido         VARCHAR(100),\n")
        w("    dni              VARCHAR(20),\n")
        w("    telefono         VARCHAR(20),\n")
        w("    correo           VARCHAR(150),\n")
        w("    fecha_nacimiento DATE,\n")
        w("    notas            TEXT\n")
        w(") ON COMMIT DROP;\n\n")

        LOTE = 500
        for i in range(0, len(registros), LOTE):
            trozo = registros[i:i + LOTE]
            w("INSERT INTO _import_pacientes (nombre, apellido, dni, telefono, correo, fecha_nacimiento, notas) VALUES\n")
            valores = []
            for (nom, ape, dni, tel, cor, fnac, notas) in trozo:
                valores.append("('%s','%s','%s',%s,%s,%s,%s)" % (
                    escapar(nom), escapar(ape), escapar(dni),
                    sql_txt(tel), sql_txt(cor),
                    ("DATE '%s'" % fnac) if fnac else "NULL",
                    sql_txt(notas)))
            w(",\n".join(valores))
            w(";\n\n")

        w("-- El Excel puede traer el mismo DNI dos veces con datos distintos: se queda el primero.\n")
        w("INSERT INTO pacientes (nombre, apellido, dni, telefono, correo, fecha_nacimiento, notas,\n")
        w("                       activo, sede_id, created_at, updated_at, saldo_a_favor)\n")
        w("SELECT DISTINCT ON (i.dni)\n")
        w("       i.nombre, i.apellido, i.dni, i.telefono, i.correo, i.fecha_nacimiento, i.notas,\n")
        w("       true,\n")
        w("       (SELECT id FROM sedes ORDER BY id LIMIT 1),\n")
        w("       now(), now(), 0\n")
        w("FROM _import_pacientes i\n")
        w("ORDER BY i.dni\n")
        w("ON CONFLICT (dni) DO NOTHING;\n\n")

        w("SELECT 'DESPUES' AS momento, count(*) AS pacientes FROM pacientes;\n")
        w("SELECT 'importados en esta corrida' AS detalle,\n")
        w("       count(*) FILTER (WHERE p.notas LIKE 'Importado del Excel del 03-09-2026%') AS filas\n")
        w("FROM pacientes p;\n\n")
        w("COMMIT;\n")

    print("Filas del Excel      : %d" % len(filas))
    print("Listas para importar : %d" % stats['listos'])
    print("  descartadas por duplicado en el Excel : %d" % stats['duplicado_en_excel'])
    print("  descartadas por basura / DNI invalido : %d" % stats['basura'])
    print("  sin DNI o sin nombre                  : %d" % stats['sin_dni_o_nombre'])
    print("Nombres marcados para revisar: %d" % len(ambiguos))
    print("\nSQL escrito en: %s" % salida)
    print("\n--- muestra del corte nombre/apellido ---")
    for r in registros[:12]:
        print("  %-10s | %-28s | %s" % (r[2], r[0], r[1]))
    print("\n--- muestra de los AMBIGUOS (conviene revisarlos) ---")
    for a in ambiguos[:12]:
        print("  %-10s %-46s -> nombre='%s'  apellido='%s'" % a)


if __name__ == '__main__':
    main()
